"""
Reporting Service - สร้างรายงานและ export ผลลัพธ์
แก้ไขให้ทำงานกับ Streamlit Cloud (in-memory export)
"""

import pandas as pd
import io
from datetime import datetime
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class ReportingService:
    """
    Service สำหรับสร้างรายงานและ export ผลลัพธ์
    """
    
    def __init__(self):
        """Initialize reporting service"""
        pass
    
    def export_to_excel_bytes(
        self, 
        reconciliation_df: pd.DataFrame,
        calculated_df: Optional[pd.DataFrame] = None,
        summary_df: Optional[pd.DataFrame] = None
    ) -> bytes:
        """
        Export ผลลัพธ์เป็น Excel ในรูปแบบ bytes (in-memory)
        สำหรับ Streamlit download_button
        
        Args:
            reconciliation_df: DataFrame ผลการเปรียบเทียบ
            calculated_df: DataFrame การคำนวณ allowances (optional)
            summary_df: DataFrame สรุปผล (optional)
            
        Returns:
            bytes: Excel file content
        """
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Sheet 1: Reconciliation (ผลการเปรียบเทียบ)
            if reconciliation_df is not None and len(reconciliation_df) > 0:
                reconciliation_df.to_excel(
                    writer, 
                    sheet_name='Reconciliation', 
                    index=False
                )
                self._format_reconciliation_sheet(
                    writer.sheets['Reconciliation'],
                    reconciliation_df
                )
            
            # Sheet 2: Calculated Allowances (การคำนวณ)
            if calculated_df is not None and len(calculated_df) > 0:
                calculated_df.to_excel(
                    writer, 
                    sheet_name='Calculated', 
                    index=False
                )
                self._format_calculated_sheet(
                    writer.sheets['Calculated'],
                    calculated_df
                )
            
            # Sheet 3: Summary (สรุปผล)
            if summary_df is not None and len(summary_df) > 0:
                summary_df.to_excel(
                    writer, 
                    sheet_name='Summary', 
                    index=False
                )
                self._format_summary_sheet(
                    writer.sheets['Summary'],
                    summary_df
                )
        
        output.seek(0)
        return output.getvalue()
    
    def _format_reconciliation_sheet(self, ws, df):
        """จัดรูปแบบ sheet Reconciliation"""
        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Column widths
        column_widths = {
            'A': 15,  # tta_key
            'B': 12,  # vendor_code
            'C': 30,  # vendor_name
            'D': 12,  # category_code
            'E': 30,  # category_name
            'F': 15,  # should_collect
            'G': 15,  # actually_collected
            'H': 15,  # difference
            'I': 12,  # status
            'J': 12,  # variance_pct
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Number formatting
        for row in range(2, len(df) + 2):
            # Format currency columns
            for col in ['F', 'G', 'H']:
                ws[f'{col}{row}'].number_format = '#,##0.00'
            
            # Format percentage
            ws[f'J{row}'].number_format = '0.00%'
            
            # Status color coding
            status_cell = ws[f'I{row}']
            status = status_cell.value
            if status and '✅' in str(status):
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif status and '❌' in str(status):
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif status and '⚠️' in str(status):
                status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        # Freeze panes
        ws.freeze_panes = 'A2'
    
    def _format_calculated_sheet(self, ws, df):
        """จัดรูปแบบ sheet Calculated"""
        # Header styling
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Column widths
        ws.column_dimensions['A'].width = 12  # vendor_code
        ws.column_dimensions['B'].width = 30  # vendor_name
        ws.column_dimensions['C'].width = 10  # division
        ws.column_dimensions['D'].width = 10  # department
        ws.column_dimensions['E'].width = 15  # tta_key
        ws.column_dimensions['F'].width = 8   # year
        ws.column_dimensions['G'].width = 15  # purchase_amount
        ws.column_dimensions['H'].width = 12  # category_code
        ws.column_dimensions['I'].width = 30  # category_name
        ws.column_dimensions['J'].width = 12  # rate_percent
        ws.column_dimensions['K'].width = 15  # fix_amount
        ws.column_dimensions['L'].width = 15  # calculated_amount
        
        # Number formatting
        for row in range(2, len(df) + 2):
            ws[f'G{row}'].number_format = '#,##0.00'  # purchase_amount
            ws[f'J{row}'].number_format = '0.00%'     # rate_percent
            ws[f'K{row}'].number_format = '#,##0.00'  # fix_amount
            ws[f'L{row}'].number_format = '#,##0.00'  # calculated_amount
        
        ws.freeze_panes = 'A2'
    
    def _format_summary_sheet(self, ws, df):
        """จัดรูปแบบ sheet Summary"""
        # Header styling
        header_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Column widths
        ws.column_dimensions['A'].width = 12  # vendor_code
        ws.column_dimensions['B'].width = 30  # vendor_name
        ws.column_dimensions['C'].width = 15  # should_collect
        ws.column_dimensions['D'].width = 15  # actually_collected
        ws.column_dimensions['E'].width = 15  # difference
        ws.column_dimensions['F'].width = 12  # status
        ws.column_dimensions['G'].width = 12  # variance_pct
        
        # Number formatting
        for row in range(2, len(df) + 2):
            ws[f'C{row}'].number_format = '#,##0.00'  # should_collect
            ws[f'D{row}'].number_format = '#,##0.00'  # actually_collected
            ws[f'E{row}'].number_format = '#,##0.00'  # difference
            ws[f'G{row}'].number_format = '0.00%'     # variance_pct
            
            # Status color coding
            status_cell = ws[f'F{row}']
            status = status_cell.value
            if status and '✅' in str(status):
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                for col in ['A', 'B', 'C', 'D', 'E', 'G']:
                    ws[f'{col}{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif status and '❌' in str(status):
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                for col in ['A', 'B', 'C', 'D', 'E', 'G']:
                    ws[f'{col}{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        ws.freeze_panes = 'A2'
    
    def export_to_csv_bytes(self, df: pd.DataFrame) -> bytes:
        """
        Export ผลลัพธ์เป็น CSV ในรูปแบบ bytes
        
        Args:
            df: DataFrame to export
            
        Returns:
            bytes: CSV file content
        """
        output = io.StringIO()
        df.to_csv(output, index=False, encoding='utf-8-sig')
        return output.getvalue().encode('utf-8-sig')
